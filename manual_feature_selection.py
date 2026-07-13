"""Excel-first manual Stage2 feature selection contract."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import numbers
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from stage2_feature_catalog import FEATURE_POOL_VERSION, feature_record, is_model_candidate


WORKBOOK_SCHEMA_VERSION = 1
FEATURE_SHEET = "Feature Selection"
SUMMARY_SHEET = "Selection Summary"
CONTRACT_SHEET = "Instructions & Contract"

SELECTION_COLUMNS = [
    "selected", "rank", "feature", "eligible", "group", "commercial_8_member",
    "commercial_original_name", "ranking_score", "train_group_fold_auc_mean",
    "valid_auc", "fp_proxy_sample_fp_rate", "valid_psi", "deployment_cost",
    "signal_source", "preprocessing", "unit", "formula", "fft", "buffer_samples",
    "accumulator", "c_operators", "risk_flags", "ineligible_reasons",
]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_ranking(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict) or not isinstance(payload.get("ranking"), list):
        raise ValueError(f"{path} must contain a ranking array")
    if payload.get("feature_pool_version") != FEATURE_POOL_VERSION:
        raise ValueError(
            f"ranking feature_pool_version must be {FEATURE_POOL_VERSION}; rerun s03-s04"
        )
    return payload


def _row_for_feature(item: dict) -> dict:
    name = str(item["feature"])
    record = feature_record(name)
    return {
        "selected": 0,
        "rank": int(item.get("rank", 0)),
        "feature": name,
        "eligible": int(bool(item.get("eligible_for_manual_selection", False))),
        "group": str(record["group"]),
        "commercial_8_member": int(bool(record["commercial_8_member"])),
        "commercial_original_name": record["commercial_original_name"] or "",
        "ranking_score": item.get("ranking_score"),
        "train_group_fold_auc_mean": item.get("train_group_fold_auc_mean"),
        "valid_auc": item.get("valid_auc"),
        "fp_proxy_sample_fp_rate": item.get("fp_proxy_sample_fp_rate"),
        "valid_psi": item.get("valid_psi"),
        "deployment_cost": float(record["deployment_cost"]),
        "signal_source": str(record["signal_source"]),
        "preprocessing": str(record["preprocessing"]),
        "unit": str(record["unit"]),
        "formula": str(record["formula"]),
        "fft": int(bool(record["fft"])),
        "buffer_samples": int(record["buffer_samples"]),
        "accumulator": str(record["accumulator"]),
        "c_operators": ", ".join(map(str, record["c_operators"])),
        "risk_flags": ", ".join(map(str, item.get("risk_flags") or [])),
        "ineligible_reasons": ", ".join(map(str, item.get("ineligible_reasons") or [])),
    }


def export_manual_selection_workbook(ranking_path, output_dir):
    ranking_path = Path(ranking_path).resolve()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = _load_ranking(ranking_path)
    rows = [_row_for_feature(item) for item in payload["ranking"]]
    ranking_sha = _sha256(ranking_path)

    workbook_path = output_dir / "manual_feature_selection.xlsx"
    csv_path = output_dir / "manual_feature_selection.csv"
    wb = Workbook()
    ws = wb.active
    ws.title = FEATURE_SHEET
    ws.append(SELECTION_COLUMNS)
    for row in rows:
        ws.append([row[column] for column in SELECTION_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    widths = {
        "A": 11, "B": 8, "C": 32, "D": 10, "E": 22, "F": 18, "G": 25,
        "Q": 65, "U": 45, "V": 30, "W": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=True)
        row[0].protection = Protection(locked=False)
    validation = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    validation.error = "selected must be 0 or 1"
    validation.errorTitle = "Invalid selection"
    ws.add_data_validation(validation)
    validation.add(f"A2:A{ws.max_row}")
    ws.conditional_formatting.add(
        f"A2:A{ws.max_row}", CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="C6EFCE"))
    )
    ws.conditional_formatting.add(
        f"D2:D{ws.max_row}", CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE"))
    )
    ws.protection.sheet = True
    ws.protection.enable()

    summary = wb.create_sheet(SUMMARY_SHEET)
    summary.append(["Metric", "Value", "Status / interpretation"])
    summary.append(["Selected features", f"=SUM('{FEATURE_SHEET}'!A2:A{ws.max_row})", "No count limit; must be non-empty"])
    summary.append(["Selected FFT features", f'=SUMPRODUCT(\'{FEATURE_SHEET}\'!A2:A{ws.max_row},\'{FEATURE_SHEET}\'!R2:R{ws.max_row})', "Engineering warning only"])
    summary.append(["Maximum selected buffer samples", f'=MAXIFS(\'{FEATURE_SHEET}\'!S2:S{ws.max_row},\'{FEATURE_SHEET}\'!A2:A{ws.max_row},1)', "Engineering warning only"])
    summary.append(["Feature-pool version", FEATURE_POOL_VERSION, "Immutable contract"])
    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 26
    summary.column_dimensions["C"].width = 44
    for cell in summary[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    summary.conditional_formatting.add(
        "B2", CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE"))
    )

    contract = wb.create_sheet(CONTRACT_SHEET)
    contract_rows = [
        ("Instruction", "Set selected=1 only in the Feature Selection sheet, then save the workbook."),
        ("workbook_schema_version", WORKBOOK_SCHEMA_VERSION),
        ("feature_pool_version", FEATURE_POOL_VERSION),
        ("ranking_source", ranking_path.name),
        ("ranking_sha256", ranking_sha),
        ("generated_utc", datetime.now(timezone.utc).isoformat()),
        ("selection_policy", "Any non-empty set of eligible governed features; engineering costs are warnings only."),
    ]
    for key, value in contract_rows:
        contract.append([key, value])
    contract.column_dimensions["A"].width = 30
    contract.column_dimensions["B"].width = 110
    for row in contract.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=True)
    contract.protection.sheet = True

    wb.save(workbook_path)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SELECTION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return {"workbook": workbook_path, "csv": csv_path}


def _contract_values(workbook) -> dict:
    sheet = workbook[CONTRACT_SHEET]
    return {
        str(sheet.cell(row=row, column=1).value): sheet.cell(row=row, column=2).value
        for row in range(1, sheet.max_row + 1)
    }


def _normalized_contract_cell(value):
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return int(value)
    return value


def _contract_cells_equal(actual, expected):
    actual = _normalized_contract_cell(actual)
    expected = _normalized_contract_cell(expected)
    if isinstance(actual, numbers.Real) and isinstance(expected, numbers.Real):
        return math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-12)
    return actual == expected


def load_manual_selection_workbook(path, ranking_path, train_columns, valid_columns):
    path = Path(path).resolve()
    ranking_path = Path(ranking_path).resolve()
    payload = _load_ranking(ranking_path)
    wb = load_workbook(path, data_only=False)
    required_sheets = {FEATURE_SHEET, SUMMARY_SHEET, CONTRACT_SHEET}
    if not required_sheets <= set(wb.sheetnames):
        raise ValueError(f"workbook missing sheets: {sorted(required_sheets - set(wb.sheetnames))}")
    contract = _contract_values(wb)
    if int(contract.get("workbook_schema_version", -1)) != WORKBOOK_SCHEMA_VERSION:
        raise ValueError("Instructions & Contract workbook_schema_version mismatch")
    if contract.get("feature_pool_version") != FEATURE_POOL_VERSION:
        raise ValueError("Instructions & Contract feature_pool_version mismatch")
    if contract.get("ranking_sha256") != _sha256(ranking_path):
        raise ValueError("Instructions & Contract ranking SHA256 mismatch")

    ranking_by_feature = {str(item["feature"]): item for item in payload["ranking"]}
    ws = wb[FEATURE_SHEET]
    headers = [str(cell.value) for cell in ws[1]]
    if headers != SELECTION_COLUMNS:
        raise ValueError(
            f"{FEATURE_SHEET} columns changed; expected immutable order {SELECTION_COLUMNS}"
        )
    index = {name: headers.index(name) + 1 for name in headers}
    expected_rows = [_row_for_feature(item) for item in payload["ranking"]]
    if ws.max_row != len(expected_rows) + 1:
        raise ValueError(
            f"{FEATURE_SHEET} row count changed; expected {len(expected_rows)} governed features"
        )
    for row_number, expected in enumerate(expected_rows, start=2):
        for field in SELECTION_COLUMNS:
            if field == "selected":
                continue
            actual_value = ws.cell(row=row_number, column=index[field]).value
            expected_value = expected[field]
            if not _contract_cells_equal(actual_value, expected_value):
                raise ValueError(
                    f"{FEATURE_SHEET} row {row_number} field {field} for "
                    f"{expected['feature']} changed immutable contract value"
                )
    selected = []
    errors = []
    for row in range(2, ws.max_row + 1):
        selected_value = ws.cell(row=row, column=index["selected"]).value
        if selected_value not in (0, 1, "0", "1", False, True):
            errors.append(f"{FEATURE_SHEET} row {row} selected must be 0 or 1")
            continue
        if int(selected_value) != 1:
            continue
        name = str(ws.cell(row=row, column=index["feature"]).value or "").strip()
        if name in selected:
            errors.append(f"{FEATURE_SHEET} row {row} duplicate feature {name}")
            continue
        item = ranking_by_feature.get(name)
        if not is_model_candidate(name) or item is None:
            errors.append(f"{FEATURE_SHEET} row {row} unknown feature {name}")
            continue
        if not bool(item.get("eligible_for_manual_selection", False)):
            errors.append(f"{FEATURE_SHEET} row {row} ineligible feature {name}")
        if name not in set(train_columns) or name not in set(valid_columns):
            errors.append(f"{FEATURE_SHEET} row {row} feature {name} missing from train/valid")
        selected.append(name)
    if errors:
        raise ValueError("manual workbook validation failed: " + "; ".join(errors))
    if not selected:
        raise ValueError("manual workbook has empty selection")

    selected_records = [feature_record(name) for name in selected]
    fft_sources = sorted({str(record["signal_source"]) for record in selected_records if record["fft"]})
    warnings = [
        f"selected_feature_count={len(selected)} (no selection limit)",
        f"fft_sources={fft_sources or []} (engineering warning only)",
        f"max_buffer_samples={max(int(record['buffer_samples']) for record in selected_records)}",
    ]
    provenance = {
        "feature_selection_mode": "manual",
        "selection_source_type": "xlsx",
        "feature_pool_version": FEATURE_POOL_VERSION,
        "manual_feature_file": str(path),
        "manual_feature_file_sha256": _sha256(path),
        "ranking_source": str(ranking_path),
        "ranking_source_sha256": _sha256(ranking_path),
        "engineering_warnings": warnings,
    }
    return selected, provenance
