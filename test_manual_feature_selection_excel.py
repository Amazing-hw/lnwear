import json
from pathlib import Path

import openpyxl
import pytest

import stage2_feature_catalog as catalog


def _ranking_payload():
    ranking = []
    for rank, name in enumerate(catalog.model_candidate_names(), start=1):
        record = catalog.feature_record(name)
        ranking.append({
            "rank": rank,
            "feature": name,
            "eligible_for_manual_selection": True,
            "ineligible_reasons": [],
            "ranking_score": 1.0 / rank,
            "group": record["group"],
            "formula": record["formula"],
            "preprocessing": record["preprocessing"],
            "unit": record["unit"],
            "valid_auc": 0.5,
            "train_group_fold_auc_mean": 0.5,
            "fp_proxy_sample_fp_rate": 0.01,
            "valid_psi": 0.0,
            "deployment_cost": record["deployment_cost"],
            "risk_flags": [],
        })
    return {
        "schema_version": 1,
        "feature_pool_version": catalog.FEATURE_POOL_VERSION,
        "ranking_policy": {"score_formula": "ranking_score"},
        "ranking": ranking,
    }


def _write_ranking(path: Path):
    path.write_text(json.dumps(_ranking_payload()), encoding="utf-8")


def test_excel_selection_export_has_complete_ordered_contract(tmp_path):
    from manual_feature_selection import export_manual_selection_workbook

    ranking_path = tmp_path / "feature_ranking_full.json"
    _write_ranking(ranking_path)

    outputs = export_manual_selection_workbook(ranking_path, tmp_path)

    assert outputs["workbook"] == tmp_path / "manual_feature_selection.xlsx"
    assert outputs["csv"] == tmp_path / "manual_feature_selection.csv"
    workbook = openpyxl.load_workbook(outputs["workbook"])
    assert workbook.sheetnames == [
        "Feature Selection",
        "Selection Summary",
        "Instructions & Contract",
    ]
    sheet = workbook["Feature Selection"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.freeze_panes == "C2"
    assert sheet.auto_filter.ref == sheet.dimensions
    assert headers[:5] == ["selected", "rank", "feature", "eligible", "group"]
    assert {"commercial_8_member", "commercial_original_name", "c_operators"} <= set(headers)
    assert sheet.max_row == len(catalog.model_candidate_names()) + 1
    assert [sheet.cell(row=i, column=3).value for i in range(2, sheet.max_row + 1)] == catalog.model_candidate_names()
    assert [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)] == [0] * len(catalog.model_candidate_names())
    assert any(validation.formula1 == '"0,1"' for validation in sheet.data_validations.dataValidation)
    assert sheet.protection.sheet is True
    assert all(sheet.cell(row=i, column=1).protection.locked is False for i in range(2, sheet.max_row + 1))


def test_excel_selection_import_preserves_any_nonempty_workbook_order(tmp_path):
    from manual_feature_selection import (
        export_manual_selection_workbook,
        load_manual_selection_workbook,
    )

    ranking_path = tmp_path / "feature_ranking_full.json"
    _write_ranking(ranking_path)
    outputs = export_manual_selection_workbook(ranking_path, tmp_path)
    workbook = openpyxl.load_workbook(outputs["workbook"])
    sheet = workbook["Feature Selection"]
    selected = ["COMM_GREEN_AC", "ACC_MAG_MEAN", "GREEN_FFT_PEAK_MEDIAN_RATIO"]
    feature_rows = {
        sheet.cell(row=row, column=3).value: row
        for row in range(2, sheet.max_row + 1)
    }
    for name in selected:
        sheet.cell(row=feature_rows[name], column=1).value = 1
    workbook.save(outputs["workbook"])

    actual, provenance = load_manual_selection_workbook(
        outputs["workbook"],
        ranking_path,
        train_columns=set(catalog.model_candidate_names()),
        valid_columns=set(catalog.model_candidate_names()),
    )

    expected_order = [name for name in catalog.model_candidate_names() if name in selected]
    assert actual == expected_order
    assert provenance["feature_selection_mode"] == "manual"
    assert provenance["selection_source_type"] == "xlsx"
    assert provenance["engineering_warnings"]


def test_excel_selection_rejects_empty_selection_and_ranking_tamper(tmp_path):
    from manual_feature_selection import (
        export_manual_selection_workbook,
        load_manual_selection_workbook,
    )

    ranking_path = tmp_path / "feature_ranking_full.json"
    _write_ranking(ranking_path)
    outputs = export_manual_selection_workbook(ranking_path, tmp_path)

    with pytest.raises(ValueError, match="empty selection"):
        load_manual_selection_workbook(
            outputs["workbook"], ranking_path,
            train_columns=set(catalog.model_candidate_names()),
            valid_columns=set(catalog.model_candidate_names()),
        )

    workbook = openpyxl.load_workbook(outputs["workbook"])
    workbook["Feature Selection"].cell(row=2, column=1).value = 1
    workbook.save(outputs["workbook"])
    payload = json.loads(ranking_path.read_text(encoding="utf-8"))
    payload["ranking"][0]["ranking_score"] = 999.0
    ranking_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ValueError, match="SHA256"):
        load_manual_selection_workbook(
            outputs["workbook"], ranking_path,
            train_columns=set(catalog.model_candidate_names()),
            valid_columns=set(catalog.model_candidate_names()),
        )


def test_excel_selection_rejects_changed_immutable_cells_even_when_row_is_unselected(tmp_path):
    from manual_feature_selection import (
        export_manual_selection_workbook,
        load_manual_selection_workbook,
    )

    ranking_path = tmp_path / "feature_ranking_full.json"
    _write_ranking(ranking_path)
    outputs = export_manual_selection_workbook(ranking_path, tmp_path)
    workbook = openpyxl.load_workbook(outputs["workbook"])
    sheet = workbook["Feature Selection"]
    headers = [cell.value for cell in sheet[1]]
    formula_col = headers.index("formula") + 1
    sheet.cell(row=2, column=formula_col).value = "tampered formula"
    sheet.cell(row=3, column=1).value = 1
    workbook.save(outputs["workbook"])

    with pytest.raises(ValueError, match=r"Feature Selection row 2 field formula.*immutable"):
        load_manual_selection_workbook(
            outputs["workbook"], ranking_path,
            train_columns=set(catalog.model_candidate_names()),
            valid_columns=set(catalog.model_candidate_names()),
        )

def test_s04_ranking_export_also_writes_excel_selection_package(tmp_path):
    import s04_feature_selection as s04

    outputs = s04.export_full_feature_ranking(
        tmp_path, _ranking_payload()["ranking"]
    )

    assert outputs["workbook"].exists()
    assert outputs["selection_csv"].exists()


def test_s05_dispatches_xlsx_selection_and_freezes_json(tmp_path):
    import pandas as pd
    import s04_feature_selection as s04
    import s05_train_final_model as s05

    ranking_outputs = s04.export_full_feature_ranking(
        tmp_path, _ranking_payload()["ranking"]
    )
    workbook = openpyxl.load_workbook(ranking_outputs["workbook"])
    sheet = workbook["Feature Selection"]
    selected = ["COMM_AMB_AC", "GREEN_CORR"]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value in selected:
            sheet.cell(row=row, column=1).value = 1
    workbook.save(ranking_outputs["workbook"])
    frame = pd.DataFrame({name: [0.0, 1.0] for name in catalog.model_candidate_names()})
    frame["target"] = [0, 1]
    frame["sample_name"] = ["a", "b"]
    frame["feature_pool_version"] = catalog.FEATURE_POOL_VERSION

    actual, provenance = s05.load_manual_feature_selection(
        ranking_outputs["workbook"], ranking_outputs["json"], frame, frame
    )

    assert actual == [name for name in catalog.model_candidate_names() if name in selected]
    frozen = json.loads(
        (tmp_path / "manual_selected_features.json").read_text(encoding="utf-8")
    )
    assert frozen["selected_features"] == actual
    assert provenance["selection_source_type"] == "xlsx"
